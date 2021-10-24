import openpyxl as excel

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "整数"
sheet["B1"] = "2進数"
sheet["C1"] = "16進数"

quantity = int(input("入力したい整数の数を決めてください："))

for i in range(quantity):
    num = int(input("整数を入力してください:"))
    sheet.cell(row=i + 2, column=1).value = num
    sheet.cell(row=i + 2, column=2).value = format(num,'b')
    sheet.cell(row=i + 2, column=3).value = format(num,'x')


file_template = "binary_excel.xlsx"
book.save(file_template)
